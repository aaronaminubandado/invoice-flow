"""
Shared export utilities for generating CSV, Excel (.xlsx), and PDF files
from tabular data.
"""

import csv
import io
from typing import List, Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def generate_csv(headers: List[str], rows: List[List[Any]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    return output.getvalue()


def generate_xlsx(headers: List[str], rows: List[List[Any]], sheet_name: str = "Export") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    alt_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(ws.cell(row=1, column=col_idx).value))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_length = max(max_length, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max_length + 4, 50
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generate_pdf_table(
    headers: List[str],
    rows: List[List[Any]],
    title: str = "Export",
) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Heading1"]))
    elements.append(Spacer(1, 20))

    table_data = [headers]
    for row in rows:
        table_data.append([str(v) if v is not None else "" for v in row])

    num_cols = len(headers)
    available_width = landscape(letter)[0] - 1.0 * inch
    col_width = available_width / num_cols

    t = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("TOPPADDING", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
                ("TOPPADDING", (0, 1), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()
